"""
Authoritative CI keying dictionaries (controlled vocabularies), loaded from the
external "Dictionary Updated" folder and used to snap free-text OCR values
(names, places, ranks, units, occupations...) to their correct canonical French
spelling as part of post-processing.

Each source .xlsx is a single column of accepted values (row 1 is a header we
skip). Values are cached to a pickle next to this file (rebuilt automatically
if the source .xlsx changes) so the ~1.74M-row surname list only has to be
parsed from Excel once.
"""
import pickle
import re
import unicodedata
from functools import lru_cache
from pathlib import Path

import openpyxl
from rapidfuzz import process, fuzz

import fields as F

DICT_DIR = Path(r"C:\Users\alien\Downloads\Dictionary Updated (4)\Dictionary Updated")
CACHE_DIR = Path(__file__).resolve().parent / "output" / "dict_cache"
CACHE_DIR.mkdir(parents=True, exist_ok=True)

# dict key -> source file(s) (Month + FRCMonth are merged into one "month" pool)
_SOURCES = {
    "city":       ["Dict_STND_France_MaineetLoireCity_Key.xlsx"],
    "state":      ["Dict_STND_France_State_Key.xlsx"],
    "eye":        ["Dict_STND_French_EyeColor_Key.xlsx"],
    "hair":       ["Dict_STND_French_HairColor_Key.xlsx"],
    "month":      ["Dict_STND_French_Month_Key.xlsx", "Dict_STND_French_FRCMonth_Key.xlsx"],
    "prefix":     ["Dict_STND_French_Prefix_Key.xlsx"],
    "suffix":     ["Dict_STND_French_Suffix_Key.xlsx"],
    "rank":       ["Dict_STND_French_Rank_Key.xlsx"],
    "regiment":   ["Dict_STND_French_MilitaryRegiment_Key.xlsx"],
    "unit":       ["Dict_STND_French_MilitaryUnit_Key.xlsx"],
    "branch":     ["Dict_STND_French_MilitaryBranch_Key.xlsx"],
    "company":    ["Dict_STND_French_MilitaryCompany_Key.xlsx"],
    "battalion":  ["Dict_STND_French_MilitaryBataillon_Key.xlsx"],
    "occupation": ["Dict_STND_French_OccupationMilitary_Key.xlsx"],
    "givenname":  ["Dict_STND_French_GivenName_Key.xlsx"],
    "surname":    ["Dict_STND_French_Surname_Key_part1.xlsx", "Dict_STND_French_Surname_Key_part2.xlsx"],
}
# large pools get bucketed by first letter so a fuzzy lookup doesn't scan
# the whole list (surname alone is ~1.74M rows)
_BIG = {"surname", "givenname", "occupation"}

# the hair/eye colour sheets mix real French vocabulary with English/French
# keying-legend glosses ("Brown, Brunet, or Marron", "Black or Noir") — those
# would corrupt a field if fuzzy-matched, so they're dropped for these keys.
_GLOSS_FILTER_KEYS = {"hair", "eye"}
_GLOSS_RE = re.compile(r",|\bor\b", re.I)

# fields.py ships small, project-curated (Aude/Gironde) canonical lists for
# these same keys — merge them in so project-specific entries (that may be
# missing from the generic dictionary, e.g. "Aude" itself) are never lost.
_MERGE_WITH_FIELDS_VOCAB = {"city": "city", "state": "state", "prefix": "prefix", "suffix": "suffix"}


def _fields_vocab_values(vocab_key: str) -> list[str]:
    spec = F.VOCAB_SETS.get(vocab_key)
    if spec is None:
        return []
    return list(spec.keys()) if isinstance(spec, dict) else list(spec)


def _norm(s: str) -> str:
    s = unicodedata.normalize("NFKD", str(s))
    s = "".join(c for c in s if not unicodedata.combining(c))
    return s.lower().strip()


def _clean(v) -> str:
    return str(v).replace("\xa0", " ").strip()


def _read_xlsx_column(path: Path) -> list[str]:
    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    out = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        v = row[0]
        if v is None:
            continue
        s = _clean(v)
        if not s or s == ".":
            continue
        out.append(s)
    wb.close()
    return out


def _build(key: str) -> dict:
    files = _SOURCES.get(key, [])
    existing = [DICT_DIR / f for f in files if (DICT_DIR / f).exists()]
    cache = CACHE_DIR / f"{key}.pkl"
    src_mtime = max((p.stat().st_mtime for p in existing), default=0)
    if src_mtime and cache.exists() and cache.stat().st_mtime >= src_mtime:
        try:
            with open(cache, "rb") as f:
                return pickle.load(f)
        except Exception:
            pass
    values, seen = [], set()
    for p in existing:
        for v in _read_xlsx_column(p):
            if key in _GLOSS_FILTER_KEYS and _GLOSS_RE.search(v):
                continue
            nk = _norm(v)
            if nk in seen:
                continue
            seen.add(nk)
            values.append(v)
    for v in _fields_vocab_values(_MERGE_WITH_FIELDS_VOCAB.get(key, "")):
        nk = _norm(v)
        if nk in seen:
            continue
        seen.add(nk)
        values.append(v)
    buckets = {}
    if key in _BIG:
        for v in values:
            nk = _norm(v)
            buckets.setdefault(nk[0] if nk else "#", []).append(v)
    data = {"values": values, "buckets": buckets, "norm_index": {_norm(v): v for v in values}}
    if values:
        try:
            with open(cache, "wb") as f:
                pickle.dump(data, f)
        except Exception:
            pass
    return data


@lru_cache(maxsize=None)
def get(key: str) -> dict:
    if key not in _SOURCES:
        return {"values": [], "buckets": {}, "norm_index": {}}
    return _build(key)


def available() -> bool:
    return DICT_DIR.exists()


def warm_all():
    """Pre-load + cache every dictionary (call once at server startup)."""
    for k in _SOURCES:
        get(k)


def stats() -> dict:
    return {k: len(get(k)["values"]) for k in _SOURCES}


def _pool_for(value_norm: str, key: str) -> list[str]:
    d = get(key)
    if key in _BIG and value_norm:
        pool = d["buckets"].get(value_norm[0])
        if pool:
            return pool
    return d["values"]


def _norm_pool(pool: list[str], query_len: int) -> dict:
    """Normalize candidates, dropping ones whose length is far from the query
    — short candidates (initials like "J", "A") otherwise get inflated WRatio
    scores against much longer misspelled queries."""
    lo = max(2, round(query_len * 0.7))
    hi = query_len + max(3, round(query_len * 0.5))
    out = {}
    for v in pool:
        nv = _norm(v)
        if lo <= len(nv) <= hi:
            out[nv] = v
    return out


def best_match(value: str, key: str, threshold: int = 84):
    """Return (matched_value, score) if `value` should be snapped to a known
    dictionary entry, or (None, 0) if it's already correct / no good match."""
    value = (value or "").strip()
    if not value:
        return None, 0
    d = get(key)
    if not d["values"]:
        return None, 0
    nk = _norm(value)
    exact = d["norm_index"].get(nk)
    if exact is not None:
        return (exact, 100) if exact != value else (None, 0)
    pool = _pool_for(nk, key)
    if not pool:
        return None, 0
    norm_pool = _norm_pool(pool, len(nk))
    if not norm_pool:
        return None, 0
    # plain ratio, not WRatio: WRatio's partial-match component badly
    # over-scores "Nth X" vs "Nth Y" style entries that share a numeric
    # prefix (e.g. "2e Bataillon" vs "2e Corps" scores 85 on WRatio but
    # only 40 on a straight ratio, which is what it should be)
    hit = process.extractOne(nk, list(norm_pool.keys()), scorer=fuzz.ratio)
    if not hit or hit[1] < threshold:
        return None, 0
    return norm_pool[hit[0]], hit[1]


def token_contains_match(value: str, key: str):
    """For tiny closed-enum dictionaries (e.g. "branch": Active/Réserve/
    Territoriale): the OCR'd value is often a longer phrase that simply
    CONTAINS the correct short canonical word ("Armée active" -> "Active").
    A whole-string ratio scores that low, so check word-containment instead.
    Only sensible for small pools — returns (canonical, 100) or (None, 0)."""
    value = (value or "").strip()
    if not value:
        return None, 0
    nv = _norm(value)
    tokens = set(nv.split())
    for canon in get(key)["values"]:
        nc = _norm(canon)
        if nc == nv or nc in tokens:
            return (canon, 100) if canon != value else (None, 0)
    return None, 0


def top_candidates(value: str, key: str, k: int = 4, threshold: int = 55) -> list[str]:
    """Top-k dictionary entries near `value` — used as grounding hints for the
    LLM spelling pass (never invents, only picks among real entries)."""
    value = (value or "").strip()
    if not value:
        return []
    d = get(key)
    if not d["values"]:
        return []
    nk = _norm(value)
    pool = _pool_for(nk, key) or d["values"]
    norm_pool = _norm_pool(pool, len(nk))
    if not norm_pool:
        return []
    hits = process.extract(nk, list(norm_pool.keys()), scorer=fuzz.ratio, limit=k)
    return [norm_pool[h[0]] for h in hits if h[1] >= threshold]
