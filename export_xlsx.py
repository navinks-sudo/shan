"""Export processed records to an .xlsx matching the required delivery
template (see d_62641.IDX.001_I3718820_images_0276_0296_78pct.xlsx): one row
per PERSON, grouped under a shared "ImageName" — row-crops split from the
same multi-person ledger page are exported under their parent page's name
(matching that reference's convention of several rows sharing one image),
not their own row-crop filename.
"""
import re
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import fields as F
import pipeline as P

HEADER_FILL = PatternFill("solid", fgColor="1F3A5F")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
THIN = Side(style="thin", color="BBBBBB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# Exact column order of the required delivery template. Most map 1:1 to our
# field labels; a few need an explicit rename (accented "Département") or
# are synthesized (ImageName, DocType).
TEMPLATE_HEADERS = [
    "ImageName", "DocType", "Enlistment City", "Enlistment Département", "Classe Year",
    "Prefix", "Given Name", "Surname", "Suffix", "Birth Day", "Birth Month", "Birth Year",
    "Birth Commune", "Birth Canton", "Birth Département", "Hair Color", "Eye Color", "Height",
    "Father Prefix", "Father Given Name", "Father Surname", "Father Suffix", "Deceased Father",
    "Mother Prefix", "Mother Given Name", "Mother Maiden Name", "Mother Surname", "Mother Suffix",
    "Deceased Mother", "Domicile", "Residence Commune", "Residence Canton",
    "Regiment", "Unit", "Branch", "Compagnie", "Battalion", "Rank",
    "Discharge Day", "Discharge Month", "Discharge Year",
    "Death Day", "Death Month", "Death Year", "Death Commune", "Occupation",
    "Entry Number", "Event Type",
]
# template header -> our internal field label, for the columns that aren't
# an exact text match
_HEADER_TO_LABEL = {"Enlistment Département": "Enlistment Departement",
                    "Birth Département": "Birth Departement"}
_SYNTHETIC = {"ImageName", "DocType"}

_ROW_RE = re.compile(r"^(?P<parent>.+)_row(?P<n>\d+)\.png$")

def _group_key(name: str):
    """(parent page name without extension, row index) — row index is 0 for
    a normal (non-split) image so it always sorts before its own rows would."""
    m = _ROW_RE.match(name)
    if m:
        return m.group("parent"), int(m.group("n"))
    return Path(name).stem, 0

def build_workbook(out_path: str, use_corrected: bool = True) -> str:
    store = P.load_results()
    # results.json accumulates every image ever processed across every folder
    # the app has ever been pointed at — scope the export to the CURRENTLY
    # loaded folder's images so switching folders doesn't leak old rows from
    # a previous, unrelated document set into the workbook.
    current = set(P.list_source_images())
    names = [n for n in store if n in current]
    names.sort(key=_group_key)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    for c, h in enumerate(TEMPLATE_HEADERS, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER

    r = 2
    for name in names:
        rec = store[name]
        src = rec.get("fields_corrected" if use_corrected else "fields_raw", {})
        parent, _ = _group_key(name)
        event_type = src.get("Event Type", "")
        for c, header in enumerate(TEMPLATE_HEADERS, 1):
            if header == "ImageName":
                value = parent
            elif header == "DocType":
                value = event_type
            else:
                value = src.get(_HEADER_TO_LABEL.get(header, header), "")
            cell = ws.cell(row=r, column=c, value=value)
            cell.font = Font(name="Arial", size=10)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = BORDER
        r += 1

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 12
    for i in range(3, len(TEMPLATE_HEADERS) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 15
    ws.freeze_panes = "C2"
    ws.row_dimensions[1].height = 30

    # legend sheet mapping label -> internal field name (kept for reference)
    ws2 = wb.create_sheet("Field Map")
    for c, h in enumerate(["Field Label", "Ancestry Field Name", "Dictionary/Vocab"], 1):
        cell = ws2.cell(row=1, column=c, value=h)
        cell.fill = HEADER_FILL; cell.font = HEADER_FONT
    for rr, (label, key, vocab) in enumerate(F.MILITARY_FIELDS, start=2):
        ws2.cell(row=rr, column=1, value=label).font = Font(name="Arial", size=10)
        ws2.cell(row=rr, column=2, value=key).font = Font(name="Arial", size=10)
        ws2.cell(row=rr, column=3, value=vocab or "").font = Font(name="Arial", size=10)
    ws2.column_dimensions["A"].width = 22
    ws2.column_dimensions["B"].width = 30
    ws2.column_dimensions["C"].width = 16

    Path(out_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path

if __name__ == "__main__":
    print(build_workbook(str(P.OUT / "Military_OCR_Output.xlsx")))
